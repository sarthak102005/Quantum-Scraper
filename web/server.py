import asyncio
import os
import sys

# Fix for Windows: Playwright requires ProactorEventLoop to spawn subprocesses.
# Uvicorn defaults to SelectorEventLoop on Windows which raises NotImplementedError.
if sys.platform == "win32":
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
import io
import json
import zipfile
from pathlib import Path
from typing import Dict, Any, Optional

from fastapi import FastAPI, HTTPException, Request, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from web.crawler_manager import WebCrawlSession

app = FastAPI(title="Quantum Scraper API", version="1.0")

allowed_origins_env = os.getenv("ALLOWED_ORIGINS", "*")
if allowed_origins_env == "*":
    allowed_origins = ["*"]
else:
    allowed_origins = [origin.strip() for origin in allowed_origins_env.split(",") if origin.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Store active session globally
active_session: Optional[WebCrawlSession] = None
session_lock = asyncio.Lock()

class CrawlRequest(BaseModel):
    seed_url: str = Field(..., description="The seed website URL to start crawling from")
    limit: int = Field(50, ge=1, le=5000, description="The maximum number of products to scrape")

@app.get("/health")
async def health() -> Dict[str, str]:
    return {"status": "ok"}

@app.post("/api/crawl")
async def start_crawl(req: CrawlRequest, background_tasks: BackgroundTasks):
    global active_session
    
    async with session_lock:
        if active_session and active_session.status == "running":
            raise HTTPException(status_code=400, detail="A crawl session is already in progress.")
        
        active_session = WebCrawlSession(req.seed_url, req.limit)
        
    # Start crawl task in the background
    background_tasks.add_task(active_session.execute_crawl)
    return {"status": "started", "message": "Scraper initialized successfully."}

@app.post("/api/crawl/stop")
async def stop_crawl():
    global active_session
    if not active_session:
        raise HTTPException(status_code=404, detail="No active crawl session to stop.")
    
    active_session.set_status("stopped")
    active_session.log("WARNING", "Crawl execution manually stopped by the user.")
    return {"status": "stopped", "message": "Crawl session stopped."}

@app.post("/api/crawl/resume")
async def resume_crawl(background_tasks: BackgroundTasks):
    global active_session
    async with session_lock:
        if not active_session:
            raise HTTPException(status_code=404, detail="No active crawl session to resume.")
        if active_session.status == "running":
            raise HTTPException(status_code=400, detail="Crawl session is already running.")
        if active_session.status not in ("stopped", "failed"):
            raise HTTPException(status_code=400, detail=f"Crawl session cannot be resumed from status '{active_session.status}'.")
        
        active_session.status = "running"
        active_session.log("SYSTEM", "Resuming crawl execution from paused point...")
        
    background_tasks.add_task(active_session.execute_crawl)
    return {"status": "started", "message": "Crawl session resumed."}


@app.get("/api/crawl/logs")
async def get_logs(request: Request):
    global active_session
    
    if not active_session:
        raise HTTPException(status_code=404, detail="No crawl session initialized.")

    async def event_generator():
        q = active_session.register_listener()
        try:
            while True:
                # Check client disconnected
                if await request.is_disconnected():
                    break
                
                try:
                    # Non-blocking fetch from listener queue
                    evt = await asyncio.wait_for(q.get(), timeout=1.0)
                    yield f"event: {evt['type']}\ndata: {json.dumps(evt['data'])}\n\n"
                    q.task_done()
                except asyncio.TimeoutError:
                    # Send keep-alive comment
                    yield ": keep-alive\n\n"
                    
        finally:
            active_session.unregister_listener(q)

    return StreamingResponse(event_generator(), media_type="text/event-stream")

@app.get("/api/crawl/export")
async def export_csv():
    csv_path = Path("outputs/products.csv")
    if not csv_path.exists():
        raise HTTPException(status_code=404, detail="No products CSV found yet. Run a crawl first.")
    
    return FileResponse(
        path=csv_path,
        media_type="text/csv",
        filename="scraped_products.csv"
    )

@app.get("/api/crawl/export/json-zip")
async def export_json_zip():
    """
    Builds an in-memory ZIP archive containing one JSON file per scraped product.
    Each JSON file is named <product_id>.json and contains the full product detail:
    specifications, description, price, image links, variants, etc.
    """
    products_dir = Path("outputs/products")
    if not products_dir.exists() or not any(products_dir.glob("*.json")):
        raise HTTPException(
            status_code=404,
            detail="No product JSON files found. Run a crawl first."
        )

    json_files = sorted(products_dir.glob("*.json"))

    # Build ZIP in memory — no temp files on disk
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for json_file in json_files:
            try:
                zf.write(json_file, arcname=json_file.name)
            except Exception:
                pass  # skip unreadable files silently
    zip_buffer.seek(0)

    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={
            "Content-Disposition": "attachment; filename=scraped_products_detail.zip"
        }
    )

@app.get("/api/crawl/export/excel")
async def export_excel():
    """
    Builds an styled Excel file (.xlsx) formatted exactly like the requested design.
    Column headers: Equipment | Category | Section | Series | Model Name | Product URL
    Segments are parsed from the product source_url.
    """
    products_dir = Path("outputs/products")
    if not products_dir.exists() or not any(products_dir.glob("*.json")):
        raise HTTPException(
            status_code=404,
            detail="No product JSON files found. Run a crawl first."
        )

    import urllib.parse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    json_files = sorted(products_dir.glob("*.json"))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Scraped Products"
    
    # Grid lines visible
    ws.views.sheetView[0].showGridLines = True

    # Setup headers
    headers = ["Equipment", "Category", "Section", "Series", "Model Name", "Product URL"]
    ws.append(headers)

    # Style definitions
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")  # Navy Blue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Alignments
    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Populate rows
    for json_file in json_files:
        try:
            with open(json_file, "r", encoding="utf-8") as f:
                p = json.load(f)
            
            source_url = p.get("source_url") or ""
            parsed = urllib.parse.urlparse(source_url)
            path_parts = [x for x in parsed.path.split("/") if x]
            
            # Clean parts: remove common locale codes / redundant parts
            clean_parts = []
            for part in path_parts:
                if part.lower() in ("en", "us", "en-us", "in", "gb", "equipment", "attachments"):
                    continue
                clean_parts.append(part)
            
            # Format helper
            def clean_name(val: str) -> str:
                # e.g., articulating-boom-lifts -> Articulating Boom Lifts
                words = val.replace("-", " ").replace("_", " ").split()
                return " ".join(w.capitalize() for w in words)

            equipment = ""
            category = ""
            section = ""
            series = ""
            model_name = p.get("title") or ""

            # Attempt parsing based on clean URL path parts depth
            if len(clean_parts) >= 1:
                equipment = clean_name(clean_parts[0])
            if len(clean_parts) >= 2:
                category = clean_name(clean_parts[1])
            if len(clean_parts) >= 3:
                section = clean_name(clean_parts[2])
            if len(clean_parts) >= 4:
                series = clean_name(clean_parts[3])
            if len(clean_parts) >= 5:
                # Override model_name if specific last path segment exists
                model_name = clean_name(clean_parts[-1])
            
            row_data = [
                equipment,
                category,
                section,
                series,
                model_name,
                source_url
            ]
            
            ws.append(row_data)
            
            # Style the data row
            current_row = ws.max_row
            for col_num in range(1, 7):
                cell = ws.cell(row=current_row, column=col_num)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = left_align
        except Exception:
            continue

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        # Add padding
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

    # Save to buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=scraped_products.xlsx"
        }
    )

# Serve Frontend static assets
static_dir = Path(__file__).parent / "static"
app.mount("/", StaticFiles(directory=str(static_dir), html=True), name="static")


